from __future__ import annotations
from pathlib import Path
from typing import List
from openpyxl import load_workbook
from .models import MiscIncome


def extract_deposits(workbook_path: Path) -> List[MiscIncome]:
    """Extract deposit-related data from company_data.xlsx"""
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    # Access the worksheet by name. wb.sheetnames is a list of names, not a mapping.
    sheet = wb["account credit nonvendor"]

    rows = sheet.iter_rows(values_only=True)
    headers = [str(h).strip() if h else "" for h in next(rows, [])]
    header_index = {h: i for i, h in enumerate(headers)}

    def _value(row, column):
        idx = header_index.get(column)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    records = []
    for row in rows:
        parent_id = _value(row, "Parent ID")
        if not parent_id:
            continue
        memo_ = _value(row, "Child ID") or ""
        amount_ = float(_value(row, "Check Amount")) or 0.0
        account_type_ = _value(row, "Tier 1 - Type") or ""
        chart_of_account_ = _value(row, "Tier 2 - Chart of Account") or ""

        records.append(
            MiscIncome(
                amount=amount_,
                memo=memo_,
                chart_of_account=chart_of_account_,
                source="excel",
            )
        )

    wb.close()
    return records


__all__ = ["extract_deposits", "MiscIncome"]


if __name__ == "__main__":  # pragma: no cover - manual invocation
    import sys

    # Allow running as a script: poetry run python src/excel_reader.py
    try:
        misc_incomes = extract_deposits(Path("company_data.xlsx"))
        for income in misc_incomes:
            print(income)
    except Exception as e:
        print(f"Error: {e}")
        print("Usage: python src/excel_reader.py <path-to-workbook.xlsx>")
        sys.exit(1)
